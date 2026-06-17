import json
from sqlite3.dbapi2 import Date

import pandas as pd
from dataclasses import dataclass
from typing import Optional
from openpyxl import load_workbook
from copy import copy

@dataclass
class BillRecord:
    收支类型: str
    账单类型: str
    金额: int
    手续费: Optional[float]
    不计收支: str
    不计预算: str
    分类: str
    子分类: str
    记账日期: str
    备注信息: str
    标签: Optional[str]
    报销状态: str
    退款状态: str
    退款金额: Optional[float]
    账本名称: Optional[str]
    账户名称: Optional[str]
    转入账户名称: Optional[str]
    借款账户名称: Optional[str]

df = pd.read_csv("asd.csv")
lzx_records: list[dict] = []
classTrans = [
    ["incomeType", "收支类型"],
    ["category", "分类"],
    ["num", "金额"],
    ["subCategory", "子分类"],
    ["date", "记账日期"],
    ["remark", "备注信息"],
]

def transferClass():
    for row in df.fillna("").to_dict(orient="records"):
        # records.append(BillRecord(**row))
        lzxRecord = {}
        for trans in classTrans:
            lzxRecord[trans[0]] = row[trans[1]]
        lzx_records.append(lzxRecord)

def readExcel():
    file_path = "target.xlsx"

    workbook = load_workbook(file_path)
    workSheet = workbook.worksheets[1]  # 第二张表（index从0开始）

    col_c = 3  # C列

    # print(workSheet.max_row)
    # print(workSheet.max_column)
    # print(workSheet.cell(row=workSheet.max_row, column=col_c).value)

    start_row = workSheet.max_row + 1

    col_date = 2
    col_cate = 3
    col_remark = 4
    col_num = 6

    def copyFormat(cell):
        prev_cell = workSheet.cell(row=cell.row - 1, column=cell.column)
        # 复制样式（关键）
        cell.font = copy(prev_cell.font)
        cell.border = copy(prev_cell.border)
        cell.fill = copy(prev_cell.fill)
        cell.number_format = copy(prev_cell.number_format)
        cell.protection = copy(prev_cell.protection)
        cell.alignment = copy(prev_cell.alignment)
        pass

    def fillDate(record, start_row):
        cell = workSheet.cell(row=start_row, column=col_date)
        date : str = record["date"]
        value = date.split(" ")[0]
        listDate = value.split("-")
        d = Date(int(listDate[0]) , int(listDate[1]), int(listDate[2]))
        fillCell(cell, d)

    def fillCategory(record, start_row):
        cell = workSheet.cell(row=start_row, column=col_cate)
        cate = record["category"]
        subCategory : str = record["subCategory"]
    #     家庭, 买菜, 皇冠, 霖, 洛, 山姆, 出去吃, 固定开销, 出去玩, 教育, 零食, 外卖
    #     家庭日常, 孩子, 洛, 教育, 买菜, 零食, 外卖
        value = subCategory
        if subCategory is None or subCategory.strip() == "":
            value = cate
        if value == '家庭日常': value = "家庭"
        # print(f"{cate} : {subCategory} : {value}")
        fillCell(cell, value)

    def fillRemark(record, start_row):
        cell = workSheet.cell(row=start_row, column=col_remark)
        remark = record["remark"]
        fillCell(cell, remark)

    def fillNum(record, start_row):
        cell = workSheet.cell(row=start_row, column=col_num)
        num : int = int(record["num"])
        fillCell(cell, num)

    def fillCell(cell, value):
        copyFormat(cell)
        cell.value = value


    for record in reversed(lzx_records):
        # print(record)
        if (record["incomeType"] != "支出"):
            continue
        fillDate(record, start_row)
        fillCategory(record, start_row)
        fillRemark(record, start_row)
        fillNum(record, start_row)
        start_row = start_row + 1

    workbook.save("result.xlsx")





    # for r in range(start_row, ws.max_row + 2):  # +2 防止最后一行也满
    #     if ws.cell(row=r, column=col_c).value is None:
    #         target_row = r
    #         break
    #
    # print("找到空行:", target_row)
    #
    # # 2. 假设你要写入的数据
    # new_data = ["A", "B", "C"]  # 示例：你可以换成自己的结构
    #
    # # 3. 写入 + 复制上一行格式
    # prev_row = target_row - 1
    #
    # for i, value in enumerate(new_data, start=1):
    #     cell = ws.cell(row=target_row, column=i)
    #     prev_cell = ws.cell(row=prev_row, column=i)
    #
    #     cell.value = value
    #
        # # 复制样式（关键）
        # cell.font = copy(prev_cell.font)
        # cell.border = copy(prev_cell.border)
        # cell.fill = copy(prev_cell.fill)
        # cell.number_format = copy(prev_cell.number_format)
        # cell.protection = copy(prev_cell.protection)
        # cell.alignment = copy(prev_cell.alignment)
    #
    # # 4. 保存
    # wb.save(file_path)

def main():
    transferClass()
    readExcel()


if __name__ == "__main__":
    main()
